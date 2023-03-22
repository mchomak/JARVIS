import openpyxl
from openpyxl.styles import Font,PatternFill
file_name=input('введите название исходного файла')
print('название исходного файла '+file_name+'.xlsx')
save_name=input('введите название сохраняемого файла')
print('название сохраняемого файла '+save_name+'.xlsx')
file=openpyxl.reader.excel.load_workbook(filename=file_name+'.xlsx')
print(file.sheetnames)
file.active=0
list1=file.active
g7=list1['G1'].value

def pol_cheker(name):
    name = name.split()
    name1 = name[0].lower()
    name2 = name[1].lower()
    if name1[-1] == 'a':
        pol = 'ЖЕН'
        print(pol)
        return pol
    elif name2[-1] == 'a':
        pol = 'ЖЕН'
        print(pol)
        return pol
    else:
        pol = 'МУЖ'
        print(pol)
        return pol

for stolbic in range(1,123):
    for acheka in range(0,7):
        print(stolbic,acheka)
        kletka = list1[stolbic][acheka].value
        print(kletka,':)')
        if kletka==None:
            if acheka==6:
                kletka2=list1['G1'].value
                list1[stolbic][acheka].value = kletka2
                kletka2=list1[stolbic][acheka]
                kletka2.font = Font(size=11)
            list1[stolbic][acheka].fill=PatternFill(start_color="FF0000", end_color="FF0000", fill_type = "solid")
        else:
            if  acheka==1:
                name=list1[stolbic][acheka].value
                pol=pol_cheker(name)
            elif acheka==2:
                list1[stolbic][acheka].value = pol
            else:
                list1[stolbic][acheka].value = kletka
        kletka = list1[stolbic][acheka].value
        print(kletka, ':)')

file.save(save_name+'.xlsx')
file.close()
