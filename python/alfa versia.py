import pyttsx3
from clock import Event
import datetime
import openpyxl
from openpyxl.styles import Font,PatternFill
from fuzzywuzzy import fuzz
import speech_recognition as sr
import time
import os
import cv2
import random
import webbrowser
import pyglet

timer_hour,timer_min=0,0
stop,day2,hour2,min2,clock_chek,timer_chek,repeat_time,time_to_stay=0,0,0,0,0,0,0,0
save1,save2,stay_timer,sit_timer,sit_minet,stay_hour=0,0,0,0,0,0
clockes,music,clock=[],[],[]
command1,command2,voice='','',''

time1 = datetime.datetime.today()
book=openpyxl.reader.excel.load_workbook(filename='clock.xlsx')
clock_musik1='D:\рамиль\PycharmProjects\J.A.R.V.I.S\sound\будильник.wav'
book.active=0
list=book.active
g7=list['G1'].value
face_cascade_db = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
eye_cascade = cv2.CascadeClassifier(cv2.data.haarcascades+"haarcascade_eye.xml")
cap = cv2.VideoCapture(0)

names=['джарвис',"жарвис","дарвис","арвис","дарвин","джанго","джаред","jahres","джара","джанис",'java']
delete=["пожалуйста","плиз","пж","ладно","эй","ой","не мог бы ты","мог бы ты","на","в","утра","дня","вечера",'эй',"хэй","ой",'сколько',"какие"]
Yes=["да","давай","ладно","одобряю","окей"]
No=['нет',"не","против","отказ"]
epic_game=['epic games','fortnite',"hitman","tabs","gta","arc",'арк']
planes={'сегодня':['today'],
        'завтра':['tomorrow'],
        'будние':["первое, питон, 2 часа","второе, ардуино 2 часа","третье, фьюжен 360, 2 часа"],
        "выходные":["поесть"],
        'каникулы':["поспать"],
}

planes_of_holidey={'первое':"питон",
                   'второе':"фьюжен 360",
                   'третье':"ардуино"
                   }
# hours=["час","часы","часов","минута","минуту","минуты","секунд","сек","секунда","секунды","секунду"]
cmd={
    'times':['покажи','текущее', "время"," час","времени"],
    'clock':['будильник'],
    'timer':["таймер"],
    'sekundomer':["секундомер"],
    'information':["запомни","запиши"],
    'poisk':["найди","загугли","поищи"],
    'translete':['переведи',"переводиться"],
    'reminder':['добавь',"событие"],
    'saund_write':["голосовой","ввод"],
    'games':["запусти"],
    'planes':["планы"],
    'vk':['vk','вк'],
    'youtube':['youtube','ютуб'],
    'google_drive':['drive','драйв','disk','диск'],
    'google':['google','гугл'],
    'school':['успеваемость','школьный'],
    'find person':['кто','такая']
}

speaking_verbs={
    'poka':['пока'],
    'privet': ["привет", "доброе", "утро", "вечер", "день"],
    'spasibo': ['спасибо', "спс", "благодарю"]

}
ru_voice_id='HKEY_LOCAL_MACHINE\SOFTWARE\Microsoft\Speech\Voices\TokenEnums\RHVoice\Aleksandr'
rech=sr.Recognizer()
micro=sr.Microphone(device_index=1)

def audio(files):
    media = pyglet.media.load(files)
    media.play()

def poisk_slova(text,slovo):
    if text.find(slovo)>=0:
        return

def poisk_slov(text,slova):
    for slovo in slova:
        if text.find(slovo) >= 0:
            return True
        else:
            pass

def Speak(what):
    print(what)
    speak_engine.say(what)
    speak_engine.runAndWait()
    speak_engine.stop()

def priv():
    time1 = datetime.datetime.today()
    audio('D:\рамиль\PycharmProjects\J.A.R.V.I.S\sound\\приветствие.wav')

def NO_or_Yes(predlog):
    predlog = predlog.split(' ')
    for slovo in predlog:
        for da in Yes:
            k = fuzz.ratio(slovo, da)
            if k >= 80:
                return True
    for slovo in predlog:
        for no in No:
            k = fuzz.ratio(slovo, no)
            if k >= 80:
                return False
    return None

def obrabotka_picture(img,faces,img_gray):
    for (x,y,w,h) in faces:
        cv2.rectangle(img, (x,y), (x+w,y+h), (0,255,0), 2)
        img_gray_face = img_gray[y:y+h,x:x+w]
        eyes = eye_cascade.detectMultiScale(img_gray_face, 1.1, 19)
        for (ex, ey, ew, eh) in eyes:
            cv2.rectangle(img, (x+ex, y+ey), (x+ex + ew, y+ey + eh), (255, 0, 0), 2)
    return img

def men_detected():
    users_time_file = open('D:\рамиль\PycharmProjects\J.A.R.V.I.S\\txt_documents\\users time.txt', 'w')
    global save1,save2,stay_timer,sit_timer,sit_minet,stay_hour,stay_minet,time_to_stay,repeat_time
    success, img = cap.read()
    img_gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    faces = face_cascade_db.detectMultiScale(img_gray, 1.1, 19)
    if faces == ():
        men = 0
        print('men not detected')
        now = datetime.datetime.now()
        if save1 == 0:
            info1 = ('stay '+str(now.year) + '.' + str(now.month) + '.' + str(now.day) + '.' + str(now.hour) + '.' + str(now.minute) + '\n')
            users_time_file.write(info1)
            save1, save2 = 1, 0
        if sit_timer == 0:
            print('sit_timer on')
            sit_minet = now.minute + 5
            sit_timer = 1
        else:
            now = datetime.datetime.now()
            if now.minute == sit_minet:
                print('sit_timer off')
                Speak('досвидание сэр')
                stay_hour,time_to_stay,stay_timer = 0,0,0
                sit_minet=-1
    else:
        men = 1
        sit_timer = 0
        print('men detected')
        now = datetime.datetime.now()
        if save2 == 0:
            info2 = ('sit '+str(now.year) + '.' + str(now.month) + '.' + str(now.day) + '.' + str(now.hour) + '.' + str(now.minute) + '\n')
            users_time_file.write(info2)
            save1, save2 = 0, 1
        if stay_timer == 0:
            print('stay_timer on')
            Speak('здравствуйте сэр')
            stay_hour = now.hour+1
            stay_minet=now.minute
            Speak('таймер прозвенит в '+str(stay_hour)+' часов и '+str(stay_minet)+' минут ')
            stay_timer = 1
        else:
            now = datetime.datetime.now()
            if time_to_stay==1:
                if repeat_time==now.minute:
                    repeat_time=now.minute+5
                    Speak('сэр прошел один час, время вставать')
            else:
                if now.hour == stay_hour and now.minute==stay_minet:
                    print('stay_timer off')
                    time_to_stay,repeat_time = 1,now.minute
                    Speak('сэр прошел один час, время вставать')
                    sit_minet=0
    users_time_file.close()
    return men,sit_minet,stay_hour

def obrabotka(text,slovar):
    for slovo in text:
        for key in slovar:
            for sl in slovar[key]:
                k=fuzz.ratio(slovo,sl)
                if k>=75:
                    cmda=key
                    return cmda

def Cmd_speaking_mode(voice,command1,command2): #voise: весь текст command1: всё без имени command2: всё кроме имени
    print('speaking_mode')
    print(command2)
    cmda=obrabotka(command1,speaking_verbs)
    print(cmda)
    if cmda=='poka':
        Cmd_poka()
    elif cmda=='privet':
        Speak('здравствуйте сэр')
    elif cmda=='spasibo':
        Speak('пожалуйста сэр')

def Cmd_find_person():
    pass

def Cmd_vk():
    Speak('я открыл вк')
    print('vk')
    webbrowser.open('https://vk.com/feed')

def Cmd_youtube():
    Speak('я открыл ютуб')
    print('youtube')
    webbrowser.open('https://www.youtube.com')

def Cmd_school():
    Speak('я открыл школьный портал')
    print('school')
    webbrowser.open('https://schools.school.mosreg.ru/marks.aspx?school=48734&index=5&tab=week&year=2021&month=3&day=25&homebasededucation=False')

def Cmd_google():
    Speak('я открыл гугл')
    print('google')
    webbrowser.open('https://www.google.ru')

def Cmd_google_drive():
    Speak('я открыл гугл диск')
    print('google_drive')
    webbrowser.open('https://drive.google.com/drive/my-drive')

def Cmd_sps():
    Speak('пожалуйста сэр')

def Cmd_planes():
    now=datetime.datetime.now()
    day1=datetime.datetime(now.year,now.month,now.day)
    print(day1.strftime("%w"))
    for slovo in command2:
        for plane in planes:
            k=fuzz.ratio(slovo,plane)
            if k>=80:
                if planes[plane]==['today']:
                    if int(day1.strftime("%w")) >= 1 and int(day1.strftime("%w")) <= 5:
                        planes[plane]=planes['будние']
                    else:
                        planes[plane] = planes['выходные']
                elif planes[plane]==['tomorrow']:
                    if int(day1.strftime("%w"))+1 >= 1 and int(day1.strftime("%w"))+1 <= 5:
                        planes[plane]=planes['будние']
                    else:
                        planes[plane] = planes['выходные']
                Speak(planes[plane])

def Cmd_games():
    game=command2
    print(game)
    game1=''
    for slovo in game:
        game1=game1+slovo+' '
    print(game1,":)")
    if game1=='':
        pass
    elif fuzz.ratio(game1,'overwatch')>=80 or fuzz.ratio(game1,'овервотч')>=80:
        os.startfile('C:\Program Files (x86)\Overwatch\\Overwatch Launcher.exe')
    elif game1=='майнкрафт' or game1=='minecraft':
        os.startfile('C:\\Users\Ramil\AppData\Roaming\.minecraft\\TLauncher.exe')
    elif game1=='браузер':
        os.startfile('C:\\Users\Ramil\AppData\Local\Programs\Opera GX\\launcher.exe')
    elif game1=='валорант' or game1=='valorant':
        os.startfile('C:\Riot Games\Riot Client\\RiotClientServices.exe')
    elif game1=='стим' or game1=='steam':
        os.startfile('C:\Program Files (x86)\Steam\\steam.exe')
    for game2 in epic_game:
        k=fuzz.ratio(game1,game2)
        if k>=80:
            os.startfile('C:\Program Files (x86)\Epic Games\Launcher\Portal\Binaries\Win32\\EpicGamesLauncher.exe')
            break

def Cmd_saund_write():
    saund_write=''
    Speak('голосовой ввод активирован')
    while True:
        Speak('слушаю')
        voice=sluh(180)
        saund_write=saund_write+str(voice)+". "
        k=fuzz.ratio(voice,'стоп')
        if k>=80:
            Speak('голосовой ввод завершен')
            break
    while True:
        Speak('мне запомнить этот текст?')
        voice=sluh(3)
        otvet=NO_or_Yes(voice)
        if otvet==True:
            file=open('D:\рамиль\PycharmProjects\J.A.R.V.I.S\\txt_documents\\saund_write.txt','w')
            file.write(saund_write)
            Speak('я сохранил текст')
            break
        elif otvet==False:
            Speak('текст не сохранен')
            break
        else:
            pass
        print(saund_write)
        time.sleep(1)


def Cmd_save_information():
    global command1,voice
    print("save_information")
    file=open('D:\рамиль\PycharmProjects\J.A.R.V.I.S\\txt_documents\\информация.txt','a')
    information=' '.join(command1)
    file.write(information + '\n')
    Speak("я записал " + information)

def Cmd_poisk():
    global command1,voice
    print("save_information")
    print(command1)
    poisk=' '.join(command1)
    print(poisk)
    webbrowser.open('https://yandex.ru/search/?clid=2358536&text='+poisk+'&l10n=ru&lr=213')
    Speak("я нашел " + poisk)

def Cmd_translate():
    global command1, voice
    print("save_information")
    translate=' '.join(command1)
    webbrowser.open('https://translate.google.ru/?sl=en&tl=ru&text='+translate+'&op=translate')
    Speak("я перевел " + translate)

def Cmd_reminder():
    buf =''
    file=open('D:\рамиль\PycharmProjects\J.A.R.V.I.S\\txt_documents\\команда.txt', 'w')
    for slovo in command1:
        buf=buf+str(slovo)+' '
    file.write(buf)
    Speak('событие создано ')
    os.startfile(r"D:\\рамиль\\PycharmProjects\\J.A.R.V.I.S\dist\\reminder.exe")

def Cmd_poka():
    print('poka')
    Speak('досвидание сэр')
    exit()

def Cmd_time():
    now = datetime.datetime.now()
    Speak("Сейчас " + str(now.hour) + ":" + str(now.minute))

def Cmd_clock(command2):
    print("clock")
    print(command2)
    file=openpyxl.Workbook()
    file=file.active()
    print(voice)
    file.write(str(voice))
    Speak("будильник поставлен")
    os.startfile(r"D:\\рамиль\\PycharmProjects\\J.A.R.V.I.S\dist\\clock.exe")

def Cmd_timer():
    print("timer")
    print(command2)
    Speak("Таймер поставлен")

def Cmd_sekundomer():
    print("sekundomer")
    Speak("Секундомер вклячён")
    os.startfile(r"D:\\рамиль\\PycharmProjects\\J.A.R.V.I.S\dist\\sekundomer.exe")




def checker():
    name, opisanie, audio, type = Event.cheker('gg')
    if name!=0:
        Speak("сэр событие "+name+' с описанием '+opisanie+" наступило ")
        Speak(audio)

def sluh(limit):
    global voice, rech, micro, time1,cmd
    print(":)")
    voice=''
    with sr.Microphone() as sourse:
        rech.adjust_for_ambient_noise(sourse)
        audio = rech.listen(sourse, phrase_time_limit=limit)
        try:
            voice = (rech.recognize_google(audio, language="ru-RU")).lower()
        except(sr.UnknownValueError):
            pass
        except(TypeError):
            pass
        print(voice)
        return voice

def deletes(command):
    global command1,command2,voice
    print("deletes")
    command1 = command.split(' ')
    print(command1)
    command1=delete1(command1)
    if command1==None:
        print("это не мне")
        return False
    else:
        command2 = delete2(command1)
        if len(command2)<=0:
            Speak("что сэр")
            return False
        else:
            return command2

def delete1(command):
    index=0
    for slovo in command:
        for name in names:
            k = fuzz.ratio(slovo, name)
            if k >= 75:
                print(index)
                command.pop(index)
                return command
        index = index + 1
    return None

def delete2(command):
    index = 0
    for slovo in command:
        for dell in delete:
            k = fuzz.ratio(slovo, dell)
            if k >= 75:
                command.pop(index)
                break
        index = index + 1
    index = 0
    return command

def recognize_cmd(command):
    global command2
    index=0
    command2=command
    print("recognize_cmd")
    print(command2)
    kolvo_poisk = 0
    for slovo in command2:
        for y in cmd:
            for slovo_cmd in cmd[y]:
                k = fuzz.ratio(slovo, slovo_cmd)
                if k >= 75:
                    kolvo_poisk = kolvo_poisk + 1
                    for x in cmd[y]:
                        try:
                            command2.pop(index)
                        except ValueError:
                            pass
                        except IndexError:
                            pass
                        index = index + 1
                if kolvo_poisk >= 1:
                    print(slovo,slovo_cmd)
                    commanda=y
                    return commanda

def execute_cmd(command):
    print("execute_cmd")
    print(command)
    if command == 'times': # сказать текущее время
        Cmd_time()
    elif command=='clock': # поставить будильник
        Cmd_clock(command2)
    elif command=='timer': # поставить таймер
        Cmd_timer()
    elif command=='sekundomer': # воставь секундомер
        Cmd_sekundomer()
    elif command=='information': # запиши информацию
        Cmd_save_information()
    elif command=='poisk': # найди что-то
        Cmd_poisk()
    elif command=='translete': # переведи сто-то
        Cmd_translate()
    elif command=='reminder': #  напоминалка
        Cmd_reminder()
    elif command=='saund_write': #  голосовой ввод
        Cmd_saund_write()
    elif command=='games': #  запускает приложения
        Cmd_games()
    elif command == 'planes':  # говорит планы
        Cmd_planes()
    elif command == 'vk':  # открывает вк
        Cmd_vk()
    elif command == 'youtube':  # открывает ютуб
        Cmd_youtube()
    elif command == 'google':  # открывает гугл
        Cmd_google()
    elif command == 'school':  # открывает ютуб
        Cmd_school()
    elif command == 'google_drive':  # открывает диск
        Cmd_google_drive()
    elif command == 'find person':  # ищет инфу про человека
        Cmd_find_person()
    else:
        Cmd_speaking_mode(voice,command1,command2)
    command = ''

def main():
    global voice,rech,micro,time1,clock_chek
    men_detected()
    sluh(6) #voice= input() sluh(5)
    cmda=deletes(voice)
    if cmda==False:
        pass
    else:
        cmda=recognize_cmd(cmda)
        execute_cmd(cmda)
    checker()


# priv()
# time.sleep(7)
print("start")
speak_engine = pyttsx3.init()
speak_engine.setProperty('voice', ru_voice_id)
Speak("слушаю сэр")

while True:
    main()